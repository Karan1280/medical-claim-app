"""
workout_builder.py
Fills Working_Out_Sheet_BASE_TEMPLATE.xlsx (Annexure-A) with claim data,
using rate_lookup.RateLookup to compute columns E-J.

Block order (per project convention - deliberately different from the EC's
section order):
  1. final_bill (Procedure/ICU/Room-rent) - kept in extraction order
  2. return_credit lines belonging to the bill overall (kept as negative rows)
  3. hc_medicine - date ascending, one row per bill
  4. medicine - date ascending, SAME-DATE bills merged into one row
  5. discharge_medicine - date ascending, one row per bill (never merged)
  6. test - grouped by ticket (bill_no+date); first item of a ticket gets
     the Sr.No and "billno/date", later items in the same ticket get a
     blank Sr.No and ",," in the Bill No/Dt. column
  7. misc - date ascending, always last
"""

from __future__ import annotations
import io
from copy import copy
from datetime import datetime
from itertools import groupby

import openpyxl
from openpyxl.utils import get_column_letter

from .rate_lookup import RateLookup, MEDICINE_DAILY_CAP

HEADER_ROW = 11
FIRST_DATA_ROW = 13  # row 12 in the base template is a "1,2,3..10" column-index helper row we skip


def _parse_date(d):
    if not d:
        return datetime(1900, 1, 1)
    try:
        return datetime.strptime(d, "%d-%m-%Y")
    except ValueError:
        return datetime(1900, 1, 1)


def _num(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


def _rate_for_item(name: str, category: str, rl: RateLookup, room_type: str | None = None):
    """Return (entry_rate, full_rate, remark, source, confidence) for a
    final_bill / misc item by trying ICU -> Room-rent -> Package ->
    Other-procedure lookups in turn."""
    n = name.upper()
    if "ICU" in n or "CCU" in n or "HDU" in n:
        m = rl.lookup_icu(name)
        if m.found:
            return m
    if "ROOM" in n or "BED" in n:
        m = rl.lookup_room_rent(room_type or "")
        if m.found:
            return m
    m = rl.lookup_package(name)
    if m.found:
        return m
    m = rl.lookup_other_procedure(name)
    if m.found:
        return m
    return None


def _build_rows(claim: dict, rl: RateLookup, nabh_level: str) -> list[dict]:
    items = claim.get("items", [])
    room_type = (claim.get("hospital", {}) or {}).get("room_type")

    def rate_value(m):
        return m.full_rate if nabh_level == "FULL" else m.entry_rate

    rows = []

    # ---- 1. final_bill (kept in extraction order) ----
    for it in [i for i in items if i["category"] == "final_bill"]:
        gross = _num(it.get("gross_amount")) or 0
        match = _rate_for_item(it["name"], "final_bill", rl, room_type)
        row = {"name": it["name"], "bill_no": it.get("bill_no"), "date": it.get("date"),
               "gross": gross}
        if match:
            e = rate_value(match) or 0
            f = max(gross - e, 0)
            row.update(e=e, f=f, g=0, h=0, i=e,
                       remark=f"Rate as per {match.source} (matched: {match.matched_name})"
                              + ("" if match.confidence >= 0.999 else " [fuzzy match - verify]"),
                       flag=match.confidence < 0.999)
        else:
            row.update(e=None, f=None, g=None, h=None, i=None,
                       remark="Rate not listed in the uploaded PGI/Annexure-I schedule - needs manual verification",
                       flag=True)
        rows.append(row)

    # ---- 2. return / credit note lines (kept negative, mirrored) ----
    for it in [i for i in items if i["category"] == "return_credit"]:
        gross = _num(it.get("gross_amount")) or 0
        rows.append({"name": it["name"], "bill_no": it.get("bill_no"), "date": it.get("date"),
                      "gross": gross, "e": gross, "f": 0, "g": 0, "h": 0, "i": gross,
                      "remark": "Return/credit note - kept as negative line", "flag": False})

    # ---- 3. hc_medicine (date ascending, fully reimbursed) ----
    hc_items = sorted([i for i in items if i["category"] == "hc_medicine"], key=lambda x: _parse_date(x.get("date")))
    for it in hc_items:
        gross = _num(it.get("gross_amount")) or 0
        rows.append({"name": it["name"], "bill_no": it.get("bill_no"), "date": it.get("date"),
                      "gross": gross, "e": gross, "f": 0, "g": 0, "h": 0, "i": gross,
                      "remark": "High-cost injection - fully reimbursed at billed cost (Para 4a-iv)",
                      "flag": False})

    # ---- 4. medicine (merge same-date bills) ----
    med_items = sorted([i for i in items if i["category"] == "medicine"], key=lambda x: _parse_date(x.get("date")))
    for date, group in groupby(med_items, key=lambda x: x.get("date")):
        group = list(group)
        bill_nos = ",".join([g.get("bill_no") or "" for g in group])
        gross = sum(_num(g.get("gross_amount")) or 0 for g in group)
        e = min(gross, MEDICINE_DAILY_CAP)
        f = max(gross - MEDICINE_DAILY_CAP, 0)
        rows.append({"name": "MEDICINE BILL", "bill_no": bill_nos, "date": date,
                      "gross": gross, "e": e, "f": f, "g": 0, "h": 0, "i": e,
                      "remark": f"Rs. {MEDICINE_DAILY_CAP}/day cap as per Para 4(a)(iii)", "flag": False})

    # ---- 5. discharge_medicine (date ascending, never merged) ----
    dm_items = sorted([i for i in items if i["category"] == "discharge_medicine"], key=lambda x: _parse_date(x.get("date")))
    for it in dm_items:
        gross = _num(it.get("gross_amount")) or 0
        e = min(gross, MEDICINE_DAILY_CAP)
        f = max(gross - MEDICINE_DAILY_CAP, 0)
        rows.append({"name": "DISCH. MEDICINE", "bill_no": it.get("bill_no"), "date": it.get("date"),
                      "gross": gross, "e": e, "f": f, "g": 0, "h": 0, "i": e,
                      "remark": f"Rs. {MEDICINE_DAILY_CAP}/day cap as per Para 4(a)(iii)", "flag": False})

    # ---- 6. test (grouped by ticket = bill_no+date) ----
    test_items = sorted([i for i in items if i["category"] == "test"], key=lambda x: _parse_date(x.get("date")))

    def ticket_key(it):
        return (it.get("bill_no"), it.get("date"))

    for ticket, group in groupby(test_items, key=ticket_key):
        group = list(group)
        for idx, it in enumerate(group):
            gross = _num(it.get("gross_amount")) or 0
            remark = rl.lookup_fixed_remark(it["name"])
            path = rl.lookup_pathology(it["name"])
            radio = rl.lookup_radiology(it["name"]) if not path.found else None
            match = path if path.found else radio
            row = {"name": it["name"],
                   "bill_no": it.get("bill_no") if idx == 0 else None,
                   "date": it.get("date") if idx == 0 else None,
                   "ditto_combined": idx > 0,
                   "gross": gross}
            if match and match.found:
                e = rate_value(match) or 0
                f = max(gross - e, 0)
                row.update(e=e, f=f, g=0, h=0, i=e,
                           remark=remark or f"{match.source} rate (matched: {match.matched_name})"
                                  + ("" if match.confidence >= 0.999 else " [fuzzy match - verify]"),
                           flag=match.confidence < 0.999)
            else:
                row.update(e=None, f=None, g=None, h=None, i=None,
                           remark=remark or "Rate not listed in the uploaded PGI/Annexure-I schedule - needs manual verification",
                           flag=remark is None)
            rows.append(row)

    # ---- 7. misc (always last, date ascending) ----
    misc_items = sorted([i for i in items if i["category"] == "misc"], key=lambda x: _parse_date(x.get("date")))
    for it in misc_items:
        gross = _num(it.get("gross_amount")) or 0
        match = _rate_for_item(it["name"], "misc", rl, room_type)
        row = {"name": it["name"], "bill_no": it.get("bill_no"), "date": it.get("date"), "gross": gross}
        if match:
            e = rate_value(match) or 0
            f = max(gross - e, 0)
            row.update(e=e, f=f, g=0, h=0, i=e, remark=f"Rate as per {match.source}", flag=match.confidence < 0.999)
        else:
            row.update(e=None, f=None, g=None, h=None, i=None,
                       remark="Rate not listed - needs manual verification (common for ambulance/blood bank items)",
                       flag=True)
        rows.append(row)

    return rows


def build_workout_xlsx(template_path: str, claim: dict, claimant_info: dict,
                        rate_lookup: RateLookup, nabh_level: str = "FULL",
                        room_type: str | None = None) -> tuple[io.BytesIO, list[str]]:
    wb = openpyxl.load_workbook(template_path)
    ws = wb.active

    admission = claim.get("admission", {})
    patient = claim.get("patient", {})
    hospital = claim.get("hospital", {})

    ws["C3"] = claimant_info.get("claimant_name") or ""
    ws["C4"] = f"RS. {claimant_info.get('basic_pay')}" if claimant_info.get("basic_pay") else "RS. (PENDING - basic pay not provided)"
    ws["C5"] = patient.get("name") or ""
    ws["C6"] = patient.get("relation_to_claimant") or ""
    ws["C7"] = hospital.get("name") or ""
    ws["C8"] = "YES" if hospital.get("govt_panel") else ("NO" if hospital.get("govt_panel") is False else "")
    period = f"{admission.get('admission_date') or ''} TO {admission.get('discharge_date') or ''}"
    ws["C10"] = period

    rows = _build_rows(claim, rate_lookup, nabh_level)

    # style template: copy formatting from an existing filled data row (13)
    style_row = FIRST_DATA_ROW
    n_rows_needed = len(rows)

    # find current TOTAL row (search column C for "TOTAL")
    total_row_idx = None
    for r in range(FIRST_DATA_ROW, ws.max_row + 1):
        if str(ws.cell(row=r, column=3).value or "").strip().upper() == "TOTAL":
            total_row_idx = r
            break
    if total_row_idx is None:
        raise RuntimeError("Could not find TOTAL row in Working Out Sheet template")

    existing_data_rows = total_row_idx - FIRST_DATA_ROW
    extra_rows_needed = max(0, n_rows_needed - existing_data_rows)
    if extra_rows_needed:
        ws.insert_rows(total_row_idx, amount=extra_rows_needed)
        # re-locate total row after insert, and copy row style down
        total_row_idx += extra_rows_needed
        for new_r in range(total_row_idx - extra_rows_needed, total_row_idx):
            for col in range(1, 11):
                src = ws.cell(row=style_row, column=col)
                dst = ws.cell(row=new_r, column=col)
                dst.font = copy(src.font)
                dst.border = copy(src.border)
                dst.fill = copy(src.fill)
                dst.alignment = copy(src.alignment)
                dst.number_format = src.number_format

    flagged = []
    sr = 1
    r = FIRST_DATA_ROW
    for row in rows:
        combined_bill_date = None
        if row.get("ditto_combined"):
            combined_bill_date = ",,"
            sr_display = None
        else:
            bill_no = row.get("bill_no") or ""
            date = row.get("date") or ""
            combined_bill_date = f"{bill_no}/{date}" if bill_no or date else ""
            sr_display = sr
            sr += 1

        ws.cell(row=r, column=1, value=sr_display)
        ws.cell(row=r, column=2, value=row["name"])
        ws.cell(row=r, column=3, value=combined_bill_date)
        ws.cell(row=r, column=4, value=round(row["gross"], 2) if row.get("gross") is not None else None)
        ws.cell(row=r, column=5, value=round(row["e"], 2) if row.get("e") is not None else None)
        ws.cell(row=r, column=6, value=round(row["f"], 2) if row.get("f") is not None else None)
        ws.cell(row=r, column=7, value=round(row["g"], 2) if row.get("g") is not None else None)
        ws.cell(row=r, column=8, value=round(row["h"], 2) if row.get("h") is not None else None)
        ws.cell(row=r, column=9, value=round(row["i"], 2) if row.get("i") is not None else None)
        ws.cell(row=r, column=10, value=row.get("remark") or "")

        if row.get("flag"):
            flagged.append(f"Row {sr_display or '(grouped)'}: {row['name']} — {row.get('remark')}")
        r += 1

    # blank out any leftover pre-existing template rows between last written row and total
    while r < total_row_idx:
        for col in range(1, 11):
            ws.cell(row=r, column=col, value=None)
        r += 1

    last_data_row = total_row_idx - 1
    for col_letter, col_idx in zip("DEFGHI", range(4, 10)):
        ws.cell(row=total_row_idx, column=col_idx,
                value=f"=SUM({col_letter}{FIRST_DATA_ROW}:{col_letter}{last_data_row})")

    provisional = len(flagged) > 0
    worked_out_row = total_row_idx + 3
    # find "Worked out for Rs." label row (search a few rows below total)
    for rr in range(total_row_idx + 1, total_row_idx + 6):
        val = str(ws.cell(row=rr, column=2).value or "")
        if "WORKED OUT FOR RS" in val.upper():
            worked_out_row = rr
            break
    total_ref = f"I{total_row_idx}"
    label = "Worked out for Rs. " + ("(PROVISIONAL - see flagged rows) " if provisional else "")
    ws.cell(row=worked_out_row, column=2,
            value=label + f"= {total_ref} only")

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out, flagged
