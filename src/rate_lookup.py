"""
rate_lookup.py
Loads the three PGI/Haryana rate-reference workbooks and exposes lookup
functions used by workout_builder.py to fill columns E-J of the Working
Out Sheet.

Data files expected in DATA_DIR (see app.py / config):
  - Fixed_Test_Remarks_REFERENCE.xlsx   (sheet "Full Remarks")
  - PGI_Rate_Quick_Reference.xlsx       (sheets: Rules_Summary, Room_Rent,
                                          ICU_Rates, Other_Procedures_No_Package,
                                          Pathology, Radiology)
  - haryana_package_rates.xlsx          (sheet "Annexure-I (Package Rates)")
"""

from __future__ import annotations
import difflib
import re
from dataclasses import dataclass
from typing import Optional

import openpyxl


def _norm(s: str) -> str:
    """Normalise a test/procedure name for fuzzy matching."""
    if s is None:
        return ""
    s = str(s).upper().strip()
    s = re.sub(r"[^A-Z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s.strip()


@dataclass
class RateMatch:
    found: bool
    entry_rate: Optional[float] = None
    full_rate: Optional[float] = None
    matched_name: Optional[str] = None
    remark: Optional[str] = None
    room_rent_extra: Optional[str] = None
    source: Optional[str] = None
    confidence: float = 0.0  # 1.0 = exact match, <1.0 = fuzzy match


class RateLookup:
    def __init__(self, fixed_remarks_path: str, pgi_quick_ref_path: str,
                 package_rates_path: str, fuzzy_threshold: float = 0.82):
        self.fuzzy_threshold = fuzzy_threshold

        # ---- Fixed test remarks ----
        wb = openpyxl.load_workbook(fixed_remarks_path, data_only=True)
        ws = wb["Full Remarks"] if "Full Remarks" in wb.sheetnames else wb.active
        self.fixed_remarks = {}  # NORM(name) -> remark text
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row and row[0]:
                self.fixed_remarks[_norm(row[0])] = row[1]

        # ---- PGI Quick Reference ----
        wb2 = openpyxl.load_workbook(pgi_quick_ref_path, data_only=True)

        self.room_rent = []  # list of dict: category, entry, full
        for row in wb2["Room_Rent"].iter_rows(min_row=2, values_only=True):
            if row and row[1]:
                self.room_rent.append({
                    "category": row[1], "norm": _norm(row[1]),
                    "entry": row[3], "full": row[4],
                })

        self.icu_rates = []
        for row in wb2["ICU_Rates"].iter_rows(min_row=2, values_only=True):
            if row and row[1]:
                self.icu_rates.append({
                    "category": row[1], "norm": _norm(row[1]),
                    "entry": row[2], "full": row[3],
                })

        self.other_procedures = []
        for row in wb2["Other_Procedures_No_Package"].iter_rows(min_row=2, values_only=True):
            if row and row[1]:
                self.other_procedures.append({
                    "name": row[1], "norm": _norm(row[1]),
                    "rate": row[2], "code_hint": row[3],
                })

        self.pathology = {}  # norm(name) -> (entry, full, original_name)
        self._path_names = []
        for row in wb2["Pathology"].iter_rows(min_row=2, values_only=True):
            if row and row[2]:
                n = _norm(row[2])
                self.pathology[n] = (row[3], row[4], row[2])
                self._path_names.append(n)

        self.radiology = {}
        self._radio_names = []
        for row in wb2["Radiology"].iter_rows(min_row=2, values_only=True):
            if row and row[2]:
                n = _norm(row[2])
                self.radiology[n] = (row[3], row[4], row[2])
                self._radio_names.append(n)

        # ---- Full Annexure-I package rates (surgery/procedure packages) ----
        wb3 = openpyxl.load_workbook(package_rates_path, data_only=True)
        pkg_sheet = "Annexure-I (Package Rates)"
        self.packages = {}
        self._pkg_names = []
        for row in wb3[pkg_sheet].iter_rows(min_row=2, values_only=True):
            # columns: SR No, Specialty Sr No, Specialty, Treatment/Procedure,
            #          Stay, Entry Rate, Full Rate, Implant/Remarks, Room Rent
            if row and row[3]:
                n = _norm(row[3])
                self.packages[n] = {
                    "specialty": row[2], "name": row[3], "stay": row[4],
                    "entry": row[5], "full": row[6],
                    "implant_remarks": row[7], "room_rent_extra": row[8],
                }
                self._pkg_names.append(n)

    # ------------------------------------------------------------------
    def _fuzzy(self, target_norm: str, candidates: list[str]) -> Optional[tuple]:
        matches = difflib.get_close_matches(target_norm, candidates, n=1,
                                             cutoff=self.fuzzy_threshold)
        if not matches:
            return None
        best = matches[0]
        score = difflib.SequenceMatcher(None, target_norm, best).ratio()
        return best, score

    # ------------------------------------------------------------------
    def lookup_fixed_remark(self, item_name: str) -> Optional[str]:
        n = _norm(item_name)
        if n in self.fixed_remarks:
            return self.fixed_remarks[n]
        m = self._fuzzy(n, list(self.fixed_remarks.keys()))
        if m:
            return self.fixed_remarks[m[0]]
        return None

    def lookup_pathology(self, item_name: str) -> RateMatch:
        n = _norm(item_name)
        if n in self.pathology:
            e, f, orig = self.pathology[n]
            return RateMatch(True, e, f, orig, source="Pathology", confidence=1.0)
        m = self._fuzzy(n, self._path_names)
        if m:
            e, f, orig = self.pathology[m[0]]
            return RateMatch(True, e, f, orig, source="Pathology", confidence=m[1])
        return RateMatch(False)

    def lookup_radiology(self, item_name: str) -> RateMatch:
        n = _norm(item_name)
        if n in self.radiology:
            e, f, orig = self.radiology[n]
            return RateMatch(True, e, f, orig, source="Radiology", confidence=1.0)
        m = self._fuzzy(n, self._radio_names)
        if m:
            e, f, orig = self.radiology[m[0]]
            return RateMatch(True, e, f, orig, source="Radiology", confidence=m[1])
        return RateMatch(False)

    def lookup_package(self, item_name: str) -> RateMatch:
        """Surgery / procedure package lookup against full Annexure-I (1340 items)."""
        n = _norm(item_name)
        if n in self.packages:
            p = self.packages[n]
            return RateMatch(True, p["entry"], p["full"], p["name"],
                              room_rent_extra=p["room_rent_extra"],
                              source=f"Annexure-I / {p['specialty']}", confidence=1.0)
        m = self._fuzzy(n, self._pkg_names)
        if m:
            p = self.packages[m[0]]
            return RateMatch(True, p["entry"], p["full"], p["name"],
                              room_rent_extra=p["room_rent_extra"],
                              source=f"Annexure-I / {p['specialty']}", confidence=m[1])
        return RateMatch(False)

    def lookup_other_procedure(self, item_name: str) -> RateMatch:
        n = _norm(item_name)
        # try code hints like (1C) / (2C) first
        code_match = re.search(r"\((\d[A-Z])\)", item_name.upper())
        if code_match:
            code = code_match.group(1)
            for p in self.other_procedures:
                if code in (p["code_hint"] or ""):
                    return RateMatch(True, p["rate"], p["rate"], p["name"],
                                      source="Other_Procedures_No_Package", confidence=1.0)
        names = [p["norm"] for p in self.other_procedures]
        m = self._fuzzy(n, names)
        if m:
            for p in self.other_procedures:
                if p["norm"] == m[0]:
                    return RateMatch(True, p["rate"], p["rate"], p["name"],
                                      source="Other_Procedures_No_Package", confidence=m[1])
        return RateMatch(False)

    def lookup_icu(self, item_name: str) -> RateMatch:
        n = _norm(item_name)
        names = [c["norm"] for c in self.icu_rates]
        if n in names:
            c = next(c for c in self.icu_rates if c["norm"] == n)
            return RateMatch(True, c["entry"], c["full"], c["category"],
                              source="ICU_Rates", confidence=1.0)
        m = self._fuzzy(n, names)
        if m:
            c = next(c for c in self.icu_rates if c["norm"] == m[0])
            return RateMatch(True, c["entry"], c["full"], c["category"],
                              source="ICU_Rates", confidence=m[1])
        return RateMatch(False)

    def lookup_room_rent(self, room_type: str) -> RateMatch:
        """room_type like 'PVT', 'SEMI-PVT', 'GENERAL'."""
        rt = (room_type or "").upper()
        is_semi = "SEMI" in rt or "TWIN" in rt
        is_private = ("PVT" in rt or "PRIVATE" in rt) and not is_semi
        if is_private:
            bucket = "private"
        elif is_semi:
            bucket = "semi"
        else:
            bucket = "general"

        for c in self.room_rent:
            norm = c["norm"]
            if bucket == "private" and "PRIVATE" in norm and "SEMI" not in norm:
                return RateMatch(True, c["entry"], c["full"], c["category"],
                                  source="Room_Rent", confidence=1.0)
            if bucket == "semi" and "SEMI" in norm:
                return RateMatch(True, c["entry"], c["full"], c["category"],
                                  source="Room_Rent", confidence=1.0)
            if bucket == "general" and "GENERAL" in norm:
                return RateMatch(True, c["entry"], c["full"], c["category"],
                                  source="Room_Rent", confidence=1.0)
        return RateMatch(False)


MEDICINE_DAILY_CAP = 1750  # Rs./day - regular + discharge medicines (Para 4a-iii)
HIGH_COST_INJECTION_THRESHOLD = 2000  # Rs./day after 30% MRP deduction (Para 4a-iv)
